# -*- coding: utf-8 -*-
"""重点医院概览数据抽取：读 TOPCORE医院DDD.xlsx → hosp_overview.json
数据源：表格汇总/TOPCORE医院DDD*.xlsx
  - top&core医院DDD      : 医院级静态信息 + DM填写的洞察/策略/行动计划
  - top&core医院客户DDD  : 客户级明细（=用户说的 Sheet2 全部客户 / Sheet3 重点打造客户）
所有可计算指标从客户级 sheet 聚合（单一事实源），医院级 sheet 仅取静态/DM文本字段。
"""
import openpyxl, json, re, os
from collections import defaultdict, Counter

import glob as _glob
_DFLT = "/Users/zhangyun/Desktop/表格汇总/TOPCORE医院DDD8月23日.xlsx"
# 云端：EAST2_DDD 指向仓库 sources/ 下的 TOPCORE医院DDD*.xlsx
if os.environ.get("EAST2_DDD"):
    SRC = os.environ["EAST2_DDD"]
else:
    _cands = sorted(_glob.glob("/Users/zhangyun/Desktop/表格汇总/TOPCORE医院DDD*.xlsx"))
    SRC = _cands[-1] if _cands else _DFLT
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "hosp_overview.json")
print("数据源:", SRC)

MONTHS = [3, 4, 5, 6, 7, 8]  # 3月-8月
STAGES = ["不管不治", "阶梯治疗", "止痛优选", "短期治痛", "治痛管理", "资深治痛管理", "治痛管理大师"]
STAGE_IDX = {s: i for i, s in enumerate(STAGES)}

def norm_concept(s):
    """观念标准化：'短期治痛（Advocator）' -> '短期治痛'；'0'/'待开发'/空 -> None"""
    if s is None:
        return None
    s = str(s).strip()
    if s in ("", "0", "None", "nan"):
        return None
    if s == "待开发":
        return None
    # 取括号/空格前的中文部分
    m = re.match(r"^([\u4e00-\u9fa5]+)", s)
    return m.group(1) if m else s

def norm_name(n):
    """医院名标准化：去空格、去总院/院区/暨…/括号，便于后续关联（当前数据已1:1，留作鲁棒）"""
    if not n:
        return ""
    n = str(n).strip()
    n = re.sub(r"[（）()]", " ", n)
    n = re.sub(r"暨.+", "", n)          # 徐州矿务集团总医院暨…
    n = re.sub(r"总院|院区", "", n)
    n = re.sub(r"\s+", "", n)
    return n

def num(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if s in ("", "None", "nan"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0

def main():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    ws_h = wb["top&core医院DDD"]
    ws_c = wb["top&core医院客户DDD"]

    # ---------- 医院级静态信息 ----------
    hrows = list(ws_h.iter_rows(values_only=True))
    hhdr = hrows[1]  # R2
    def hcol(n):
        for i, h in enumerate(hhdr):
            if h and str(h).strip() == n:
                return i
        return None
    H = {n: hcol(n) for n in [
        "医院", "医院级别", "DM", "MICS", "准入标签", "是否有进货金额限制", "限制金额",
        "院内渠道开放情况", "梯队标签", "9月TTH目标",
        "洞察诊断", "策略", "针对各策略下客户增长及增长机会的具体行动计划", "所需需求&支持"]}
    hosp_meta = {}
    for r in hrows[2:]:
        name = r[H["医院"]]
        if name in (None, "", "医院"):
            continue
        hosp_meta[name] = {
            "level": r[H["医院级别"]] or "",
            "dm": r[H["DM"]] or "",
            "mics": r[H["MICS"]] or "",
            "access": r[H["准入标签"]] or "",
            "money_limit": r[H["是否有进货金额限制"]] or "",
            "money_limit_amt": r[H["限制金额"]] or "",
            "channel": r[H["院内渠道开放情况"]] or "",
            "tier": r[H["梯队标签"]] or "",
            "tth9_target": num(r[H["9月TTH目标"]]),
            "insight": r[H["洞察诊断"]] or "",
            "strategy": r[H["策略"]] or "",
            "action": r[H["针对各策略下客户增长及增长机会的具体行动计划"]] or "",
            "support": r[H["所需需求&支持"]] or "",
        }

    # ---------- 客户级明细 ----------
    crows = list(ws_c.iter_rows(values_only=True))
    chdr = crows[2]  # R3
    def ccol(n):
        for i, h in enumerate(chdr):
            if h and str(h).strip() == n:
                return i
        return None
    C = {n: ccol(n) for n in [
        "医院", "医院级别", "客户名", "科室", "亚专业", "客户分型", "是否list", "244标签",
        "出诊简介中是否含有头痛/疼痛、头晕/眩晕等标签", "专诊标签", "专诊次数/月", "非专诊次数/月",
        "出诊次数/月", "3月tth预估实际", "4月tth预估实际", "5月tth预估实际", "6月tth预估实际",
        "7月tth预估实际", "8月tth预估实际", "月销量变化", "现阶段观念评估（7月）", "观念提升目标",
        "9月预估目标（2片装）", "机会点", "针对每一个机会点的行动计划", "所需需求&支持"]}
    MT = {m: C[f"{m}月tth预估实际"] for m in MONTHS}

    # 按医院聚合
    hosp_customers = defaultdict(list)
    for r in crows[3:]:
        hn = r[C["医院"]]
        if hn in (None, "", "医院"):
            continue
        cust = {
            "name": r[C["客户名"]] or "",
            "dept": r[C["科室"]] or "",
            "sub": r[C["亚专业"]] or "",
            "type": (r[C["客户分型"]] or "").strip(),
            "is_list": (r[C["是否list"]] or "") == "是",
            "is244": (r[C["244标签"]] or "") == "是",
            "headache_tag": (r[C["出诊简介中是否含有头痛/疼痛、头晕/眩晕等标签"]] or "") == "是",
            "special_label": r[C["专诊标签"]] or "",
            "special_per": num(r[C["专诊次数/月"]]),
            "nonspecial_per": num(r[C["非专诊次数/月"]]),
            "visit_per": num(r[C["出诊次数/月"]]),
            "tth": {m: num(r[MT[m]]) for m in MONTHS},
            "cur_concept": norm_concept(r[C["现阶段观念评估（7月）"]]),
            "tgt_concept": norm_concept(r[C["观念提升目标"]]),
            "m9_target": num(r[C["9月预估目标（2片装）"]]),
            "opp": (r[C["机会点"]] or "").strip(),
            "action": (r[C["针对每一个机会点的行动计划"]] or "").strip(),
            "support": (r[C["所需需求&支持"]] or "").strip(),
        }
        cust["cur_idx"] = STAGE_IDX.get(cust["cur_concept"]) if cust["cur_concept"] else None
        cust["tgt_idx"] = STAGE_IDX.get(cust["tgt_concept"]) if cust["tgt_concept"] else None
        cust["tth37"] = sum(cust["tth"][m] for m in [3, 4, 5, 6, 7])
        m3, m7 = cust["tth"][3], cust["tth"][7]
        cust["growth37"] = (m7 - m3) / m3 * 100 if m3 > 0 else (None if m7 == 0 else 100.0)
        cust["is_key"] = (cust["tgt_idx"] is not None and cust["tgt_idx"] >= 4)  # 治痛管理及以上=重点打造
        hosp_customers[hn].append(cust)

    # ---------- 按医院计算指标 ----------
    hospitals = []
    for hn, cs in hosp_customers.items():
        meta = hosp_meta.get(hn, {})
        # 月度 TTH（客户求和，单一事实源）
        monthly_tth = {m: round(sum(c["tth"][m] for c in cs), 1) for m in MONTHS}
        # 月度活跃（该月 TTH>0 去重客户数）
        monthly_active = {m: sum(1 for c in cs if c["tth"][m] > 0) for m in MONTHS}
        latest = 8 if monthly_tth[8] > 0 else 7
        first_m, last_m = 3, 7
        tth_first, tth_last = monthly_tth[first_m], monthly_tth[last_m]
        net37 = round(tth_last - tth_first, 1)
        growth37 = round((tth_last - tth_first) / tth_first * 100, 1) if tth_first > 0 else None
        # 客户计数
        tag_n = sum(1 for c in cs if c["is244"])
        active_latest = monthly_active[latest]
        active_rate = round(active_latest / tag_n * 100, 1) if tag_n else 0.0
        abcd = Counter(c["type"] for c in cs if c["type"] in ("A", "B", "C", "D"))
        big_n = abcd.get("A", 0)
        total_tth7 = sum(c["tth"][7] for c in cs)
        big_tth7 = sum(c["tth"][7] for c in cs if c["type"] == "A")
        big_contrib = round(big_tth7 / total_tth7 * 100, 1) if total_tth7 else 0.0
        # 观念分布
        cur_dist = {s: 0 for s in STAGES}
        tgt_dist = {s: 0 for s in STAGES}
        for c in cs:
            if c["cur_idx"] is not None:
                cur_dist[STAGES[c["cur_idx"]]] += 1
            if c["tgt_idx"] is not None:
                tgt_dist[STAGES[c["tgt_idx"]]] += 1
        # 升级统计
        upgrade_set = sum(1 for c in cs if c["tgt_idx"] is not None and c["tgt_idx"] != c["cur_idx"])
        upgrade_planned = sum(1 for c in cs if c["tgt_idx"] is not None and c["cur_idx"] is not None and c["tgt_idx"] > c["cur_idx"])
        upgrade_none = sum(1 for c in cs if c["tgt_idx"] is None or c["tgt_idx"] == c["cur_idx"])
        upgrade_done = sum(1 for c in cs if c["tgt_idx"] is not None and c["cur_idx"] is not None and c["cur_idx"] >= c["tgt_idx"])
        has_tgt = sum(1 for c in cs if c["tgt_idx"] is not None and c["tgt_idx"] != c["cur_idx"])
        upgrade_done_rate = round(upgrade_done / has_tgt * 100, 1) if has_tgt else 0.0
        # 科室/亚专业
        dept = defaultdict(lambda: {"n": 0, "tth": 0.0})
        sub = defaultdict(lambda: {"n": 0, "tth": 0.0})
        for c in cs:
            dept[c["dept"]]["n"] += 1
            dept[c["dept"]]["tth"] += c["tth"][7]
            sub[c["sub"]]["n"] += 1
            sub[c["sub"]]["tth"] += c["tth"][7]
        # 专诊
        has_special = any(c["special_label"] not in ("", "无专诊设置") for c in cs)
        special_sum = sum(c["special_per"] for c in cs)
        nonspecial_sum = sum(c["nonspecial_per"] for c in cs)
        visit_sum = sum(c["visit_per"] for c in cs)
        headache_n = sum(1 for c in cs if c["headache_tag"])
        vertigo_n = sum(1 for c in cs if "眩晕" in (c["sub"] or ""))
        # 重点打造客户
        key_cs = [c for c in cs if c["is_key"]]
        key_tth7 = sum(c["tth"][7] for c in key_cs)
        key_contrib = round(key_tth7 / total_tth7 * 100, 1) if total_tth7 else 0.0
        # 趋势状态
        series = [monthly_tth[m] for m in MONTHS if monthly_tth[m] > 0]
        trend = trend_status(series)

        hospitals.append({
            "name": hn,
            "key": norm_name(hn),
            "level": meta.get("level", ""),
            "dm": meta.get("dm", ""),
            "mics": meta.get("mics", ""),
            "access": meta.get("access", ""),
            "money_limit": meta.get("money_limit", ""),
            "money_limit_amt": meta.get("money_limit_amt", ""),
            "channel": meta.get("channel", ""),
            "tier": meta.get("tier", ""),
            "tth9_target": meta.get("tth9_target", 0),
            "insight": meta.get("insight", ""),
            "strategy": meta.get("strategy", ""),
            "action": meta.get("action", ""),
            "support": meta.get("support", ""),
            "monthly_tth": monthly_tth,
            "monthly_active": monthly_active,
            "latest_month": latest,
            "net37": net37,
            "growth37": growth37,
            "tag_n": tag_n,
            "active_latest": active_latest,
            "target_active": None,
            "active_rate": active_rate,
            "big_n": big_n,
            "big_contrib": big_contrib,
            "abcd": {k: abcd.get(k, 0) for k in ("A", "B", "C", "D")},
            "abcd_contrib": {k: round(sum(c["tth"][7] for c in cs if c["type"] == k) / total_tth7 * 100, 1) if total_tth7 else 0.0 for k in ("A", "B", "C", "D")},
            "cur_dist": cur_dist,
            "tgt_dist": tgt_dist,
            "upgrade_planned": upgrade_planned,
            "upgrade_set": upgrade_set,
            "upgrade_none": upgrade_none,
            "upgrade_done_rate": upgrade_done_rate,
            "dept": {k: v for k, v in dept.items()},
            "sub": {k: v for k, v in sub.items()},
            "has_special": has_special,
            "special_sum": special_sum,
            "nonspecial_sum": nonspecial_sum,
            "visit_sum": visit_sum,
            "headache_n": headache_n,
            "vertigo_n": vertigo_n,
            "customers": cs,
            "key_customers": key_cs,
            "key_n": len(key_cs),
            "key_tth7": round(key_tth7, 1),
            "key_contrib": key_contrib,
            "trend_status": trend,
        })

    # 筛选维度枚举
    filters = {
        "level": sorted({h["level"] for h in hospitals if h["level"]}),
        "dm": sorted({h["dm"] for h in hospitals if h["dm"]}),
        "mics": sorted({h["mics"] for h in hospitals if h["mics"]}),
        "access": sorted({h["access"] for h in hospitals if h["access"]}),
    }
    out = {
        "meta": {"stages": STAGES, "months": MONTHS, "source": "TOPCORE医院DDD8月23日.xlsx"},
        "filters": filters,
        "hospitals": hospitals,
    }
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False)
    print("✅ 写出", OUT)
    print("   医院数:", len(hospitals), " 客户总数:", sum(len(h["customers"]) for h in hospitals))
    print("   筛选维度: 级别", filters["level"], "| DM", filters["dm"], "| 准入", filters["access"][:6], "…")
    # 抽样校验
    h0 = hospitals[0]
    print("\n样本医院:", h0["name"], h0["level"], "DM", h0["dm"])
    print("   月度TTH:", h0["monthly_tth"])
    print("   月度活跃:", h0["monthly_active"])
    print("   标签客户", h0["tag_n"], "活跃", h0["active_latest"], "活跃率", h0["active_rate"], "%")
    print("   A/B/C/D:", h0["abcd"], "大客户(A)贡献", h0["big_contrib"], "%")
    print("   观念当前:", h0["cur_dist"])
    print("   重点打造客户:", h0["key_n"], "贡献", h0["key_contrib"], "%")

def trend_status(series):
    """根据月度TTH序列判定趋势状态标签"""
    if not series or len(series) < 2:
        return "低位停滞"
    last = series[-1]
    mean = sum(series) / len(series)
    # 波动幅度
    amp = (max(series) - min(series)) / mean * 100 if mean else 0
    # 连续方向
    diffs = [series[i] - series[i - 1] for i in range(1, len(series))]
    up_streak = 0
    down_streak = 0
    for d in diffs:
        if d > 0:
            up_streak += 1; down_streak = 0
        elif d < 0:
            down_streak += 1; up_streak = 0
        else:
            up_streak = down_streak = 0
    if down_streak >= 2:
        return "连续下滑"
    if up_streak >= 2:
        return "持续增长"
    # 先降后升
    if diffs[0] < 0 and diffs[-1] > 0:
        return "恢复增长"
    if amp > 25:
        return "波动较大"
    if mean < 300:
        return "低位停滞"
    return "高位稳定"

if __name__ == "__main__":
    main()
