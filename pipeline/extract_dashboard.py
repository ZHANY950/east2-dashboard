# -*- coding: utf-8 -*-
"""提取三个数据源 → dashboard JSON（只读，不修改任何源文件）
v3（2026-08-30）：月表/DDD 路径改为 glob 最新；医院派生指标跟随最新有数据月份
"""
import json, re, glob, os
from collections import defaultdict
import openpyxl

# ===== 路径配置（环境变量可覆盖：GitHub Actions 云端用仓库 sources/ 目录，本地默认桌面） =====
SRC_DIR = os.environ.get("EAST2_SRC", "/tmp")
def _p(name): return os.path.join(SRC_DIR, name)
YB_DIRS = [d for d in os.environ.get(
    "EAST2_YB_DIR",
    "/Users/zhangyun/Desktop/基础数据源" + os.pathsep + "/Users/zhangyun/Desktop/会议有效性ppt"
).split(os.pathsep) if d]
DDD_DIRS = [d for d in os.environ.get(
    "EAST2_DDD_DIR", "/Users/zhangyun/Desktop/表格汇总").split(os.pathsep) if d]

def _pick_latest(dirs, pattern):
    """跨目录 glob 后按文件名内日期(m.d)排序取最新（'10.1'<'8.26' 字符串排序会错序）"""
    cands = []
    for d in dirs:
        cands += glob.glob(os.path.join(d, pattern))
    def key(p):
        m = re.search(r"(\d+)\.(\d+)", os.path.basename(p))
        return (int(m.group(1)), int(m.group(2))) if m else (0, 0)
    return sorted(cands, key=key)[-1] if cands else None

MONTHS = ["3月","4月","5月","6月","7月","8月"]
# DDD名单医院名 -> 月表医院名（标准化；None 表示需合并多个院区）
H_MAP = {
    "苏州大学附属第一医院": "苏州大学附属第一医院总院",
    "苏州市立医院": None,  # 合并 太湖总院+道前院区
}
H_MERGE_SUB = {"苏州市立医院": ["苏州市立医院太湖总院", "苏州市立医院道前院区"]}
# 反向映射（月表/大客户名单医院名 -> DDD名单医院名）
REV_H_MAP = {
    "苏州大学附属第一医院总院": "苏州大学附属第一医院",
    "苏州市立医院太湖总院": "苏州市立医院",
    "苏州市立医院道前院区": "苏州市立医院",
}

# 观念词表（顺序=层级）
CONCEPTS = ["不管不治","阶梯治疗","止痛优选","短期治痛","治痛管理","资深治痛管理","治痛管理大师"]
# 区域规则阈值（纯销盒数）
THRESH = [0,1,30,70,130,300,500]

def concept_from_tth(v):
    v = float(v or 0)
    for i in range(len(THRESH)-1, -1, -1):
        if v >= THRESH[i]:
            return CONCEPTS[i]
    return CONCEPTS[0]

def concept_main(text):
    """从观念文本提取主观念词，如 '治痛管理大师（Super Ambassador）' → '治痛管理大师'
    按词长降序匹配，避免'治痛管理'先命中'治痛管理大师'"""
    if not text: return ""
    t = str(text)
    for c in sorted(CONCEPTS, key=len, reverse=True):
        if c in t:
            return c
    return t.strip()[:12]

def num(v):
    try:
        return float(v) if v not in (None, "") else None
    except (TypeError, ValueError):
        return None

# ============ 1. 月表（glob 最新：基础数据源优先，会议有效性ppt备选；云端=仓库 sources/） ============
_yb = _pick_latest(YB_DIRS, "【月表】患者画像每日跟进*.xlsx")
if not _yb:
    raise SystemExit("未找到月表文件：基础数据源/会议有效性ppt/sources 均无【月表】患者画像每日跟进*.xlsx")
path_yb = _yb
print("月表使用:", path_yb)
wb = openpyxl.load_workbook(path_yb, read_only=True, data_only=True)
ws = wb["基础数据"]
# 列(0-based): 0月份 6DM 9MICS 15医院 16客户 17客户类型 21科室 22预估
y_rows = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r[0] or not r[15]: continue
    v = num(r[22])
    if v is None: continue
    y_rows.append({
        "m": str(r[0]), "dm": str(r[6]).strip(), "mics": str(r[9]).strip(),
        "h": str(r[15]).strip(), "c": str(r[16]).strip(),
        "tp": str(r[17]).strip() if r[17] else "", "v": v,
    })
wb.close()

def norm_hosp(h):
    """DDD名单医院名 -> 月表医院名集合"""
    if h in H_MAP:
        t = H_MAP[h]
        return [t] if t else H_MERGE_SUB[h]
    return [h]

# 大区聚合
month_tot = defaultdict(float)
dm_month = defaultdict(lambda: defaultdict(float))
dm_mics = defaultdict(set)
type_month = defaultdict(lambda: defaultdict(float))      # tp -> m -> tth
type_cnt_month = defaultdict(lambda: defaultdict(set))    # tp -> m -> 客户(去重)
active_month = defaultdict(set)      # m -> 客户(医院|客户)
dm_active_month = defaultdict(lambda: defaultdict(set))  # dm -> m -> 客户
mics_list = set()
for r in y_rows:
    month_tot[r["m"]] += r["v"]
    dm_month[r["dm"]][r["m"]] += r["v"]
    dm_mics[r["dm"]].add(r["mics"])
    type_month[r["tp"]][r["m"]] += r["v"]
    key = r["h"] + "|" + r["c"]
    type_cnt_month[r["tp"]][r["m"]].add(key)
    active_month[r["m"]].add(key)
    dm_active_month[r["dm"]][r["m"]].add(key)
    mics_list.add(r["mics"])

# 7月客户级观念（区域规则，纯销）
concept_dist = defaultdict(lambda: {"n":0, "tth":0.0})
concept_seen = set()
for r in y_rows:
    if r["m"] != "7月": continue
    key = r["h"] + "|" + r["c"]
    if key in concept_seen: continue
    concept_seen.add(key)
    # 该客户7月合计
    tot = sum(x["v"] for x in y_rows if x["m"]=="7月" and x["h"]+ "|"+x["c"] == key)
    c = concept_from_tth(tot)
    concept_dist[c]["n"] += 1
    concept_dist[c]["tth"] += tot

region = {
    "month_total": {m: round(month_tot.get(m,0),1) for m in MONTHS if month_tot.get(m)},
    "dm_month": {dm: {m: round(v,1) for m,v in items.items()} for dm,items in sorted(dm_month.items())},
    "dm_active": {dm: {m: len(s) for m,s in d.items()} for dm,d in dm_active_month.items()},
    "dm_mics": {dm: sorted(s) for dm,s in dm_mics.items()},
    "type_month": {tp: {m: round(v,1) for m,v in d.items()} for tp,d in type_month.items()},
    "type_cnt": {tp: {m: len(s) for m,s in d.items()} for tp,d in type_cnt_month.items()},
    "concept_dist": {c: {"n": concept_dist[c]["n"], "tth": round(concept_dist[c]["tth"],1)} for c in CONCEPTS},
    "active_month": {m: len(s) for m,s in active_month.items()},
    "mics_list": sorted(mics_list),
    "dm_list": sorted(dm_month.keys()),
}

# ============ 2. 重点医院名单（glob 最新 TOPCORE医院DDD*.xlsx；云端=仓库 sources/） ============
_ddd = _pick_latest(DDD_DIRS, "TOPCORE医院DDD*.xlsx")
if not _ddd:
    raise SystemExit("未找到重点医院名单：表格汇总/sources 均无 TOPCORE医院DDD*.xlsx")
path_ddd = _ddd
print("DDD名单使用:", path_ddd)
wb = openpyxl.load_workbook(path_ddd, read_only=True, data_only=True)

def _pick_sheet(names):
    """按候选名找 sheet（新版文件 sheet 已改名为 top&core医院DDD 等）"""
    for n in names:
        if n in wb.sheetnames:
            return wb[n]
    raise KeyError(f"sheet 不存在：{names}（现有：{wb.sheetnames}）")

ws_h = _pick_sheet(["重点医院名单", "top&core医院DDD"])
hospitals = []
for r in ws_h.iter_rows(min_row=3, values_only=True):
    if not r[0]: continue
    hosp = {
        "name": str(r[0]).strip(), "level": str(r[1]).strip() if r[1] else "",
        "dm": str(r[2]).strip() if r[2] else "", "mics": str(r[3]).strip() if r[3] else "",
        "access": str(r[4]).strip() if r[4] else "", "limit": str(r[5]).strip() if r[5] else "",
        "limitAmt": num(r[6]), "channel": str(r[7]).strip() if r[7] else "",
        "tth": {m: num(r[8+i]) for i,m in enumerate(MONTHS)},
        "tier": str(r[14]).strip() if r[14] else "", "chg": str(r[15]).strip() if r[15] else "",
        "headache_clinic": num(r[16]), "headache_cust": num(r[17]),
        "vertigo_clinic": num(r[18]), "vertigo_cust": num(r[19]),
        "tag_count": num(r[20]), "target_active": num(r[21]), "active": num(r[22]),
        "abcd": {"A": num(r[23]), "B": num(r[24]), "C": num(r[25]), "D": num(r[26])},
        "abcd_share": {"A": num(r[27]), "B": num(r[28]), "C": num(r[29]), "D": num(r[30])},
        # 列31-36 = 阶梯治疗..治痛管理大师（不含"不管不治"）
        "concept": {CONCEPTS[i+1]: num(r[31+i]) for i in range(6)},
        "q4_tier": str(r[37]).strip() if r[37] else "", "tgt9": num(r[38]),
        "insight": str(r[43]).strip() if r[43] else "", "strategy": str(r[44]).strip() if r[44] else "",
        "plan": str(r[45]).strip() if r[45] else "", "support": str(r[46]).strip() if r[46] else "",
    }
    hospitals.append(hosp)

# ============ 3. 客户名单 ============
ws_c = _pick_sheet(["客户名单", "top&core医院客户DDD"])
customers = []
for r in ws_c.iter_rows(min_row=4, values_only=True):
    if not r[0] or not r[2]: continue
    customers.append({
        "h": str(r[0]).strip(), "lv": str(r[1]).strip() if r[1] else "",
        "n": str(r[2]).strip(), "dept": str(r[3]).strip() if r[3] else "",
        "sub": str(r[4]).strip() if r[4] else "", "tp": str(r[5]).strip() if r[5] else "",
        "list": str(r[6]).strip() if r[6] else "", "t244": str(r[7]).strip() if r[7] else "",
        "clinic_desc": str(r[8]).strip() if r[8] else "", "sc": str(r[9]).strip() if r[9] else "",
        "st": num(r[10]), "nt": num(r[11]), "ct": num(r[12]),
        "tth": {m: num(r[13+i]) for i,m in enumerate(MONTHS[:6])},
        "chg": str(r[19]).strip() if r[19] else "",
        "c_now": concept_main(r[20]), "c_tgt": concept_main(r[21]),
        "tgt9": num(r[22]), "opp": str(r[23]).strip() if r[23] else "",
        "act": str(r[24]).strip() if r[24] else "", "sup": str(r[25]).strip() if r[25] else "",
    })

# ============ 4. 打造大客户名单（v3：sheet 若已移除则继承旧缓存，数据不丢） ============
key_customers = []
if "打造大客户名单" in wb.sheetnames:
    ws_k = wb["打造大客户名单"]
    for r in ws_k.iter_rows(min_row=3, values_only=True):
        if not r[3]: continue
        key_customers.append({
            "dm": str(r[1]).strip() if r[1] else "", "mics": str(r[2]).strip() if r[2] else "",
            "n": str(r[3]).strip(), "h": REV_H_MAP.get(str(r[4]).strip(), str(r[4]).strip()),
            "dept": str(r[5]).strip() if r[5] else "", "sub": str(r[6]).strip() if r[6] else "",
            "tth": {m: num(r[7+i]) for i,m in enumerate(MONTHS[:5])},
            "tgt8": num(r[12]), "chg": str(r[13]).strip() if r[13] else "",
            "c_now": concept_main(r[14]), "c_tgt": concept_main(r[15]),
            "opp": str(r[16]).strip() if r[16] else "", "act": str(r[17]).strip() if r[17] else "",
            "sup": str(r[18]).strip() if r[18] else "",
        })
    print(f"打造大客户名单: {len(key_customers)} 人（DDD文件）")
else:
    try:
        _old = json.load(open(_p("dashboard_data.json"), encoding="utf-8"))
        key_customers = _old.get("key_customers", [])
        print(f"打造大客户名单: sheet 已移除，继承旧缓存 {len(key_customers)} 人")
    except Exception:
        print("打造大客户名单: sheet 已移除且无旧缓存，置空")
wb.close()

# ============ 医院派生指标（v3：跟随最新有数据月份，缺省则用月表/客户名单补） ============
# 客户名单按医院聚合（标准化后）作为兜底 & 客户级数据
cust_by_h = defaultdict(list)
for c in customers:
    cust_by_h[c["h"]].append(c)

def _latest_idx(tth):
    """按 MONTHS 顺序取最后一个有值月份的索引；无数据返回 -1"""
    for i in range(len(MONTHS)-1, -1, -1):
        if tth.get(MONTHS[i]) is not None:
            return i
    return -1

for h in hospitals:
    li = _latest_idx(h["tth"])
    if li >= 1:
        t = [h["tth"].get(m) for m in MONTHS[:li+1]]   # 3月 .. 最新月
        vals = [v for v in t if v is not None]
        t3, tN = t[0], t[li]
        if t3 is None: t3 = vals[0]
        if tN is None: tN = vals[-1]
        h["latest_month"] = MONTHS[li]
        h["net"] = round(tN - t3, 1)
        h["growth"] = round((tN - t3) / t3 * 100, 1) if t3 else None
        mx, mn = max(vals), min(vals)
        h["max_month"] = MONTHS[:li+1][t.index(mx)]
        h["min_month"] = MONTHS[:li+1][t.index(mn)]
        avg = sum(vals)/len(vals)
        h["volatility"] = round((mx-mn)/avg*100, 1) if avg else None
        # 环比符号（首月+1 .. 最新月 vs 前月）
        signs = []
        for i in range(1, li+1):
            a, b = t[i-1], t[i]
            if a is not None and b is not None:
                signs.append(1 if b > a*1.005 else (-1 if b < a*0.995 else 0))
        up, down = signs.count(1), signs.count(-1)
        last_up = signs[-1] == 1 if signs else False
        last_down = signs[-1] == -1 if signs else False
        # 趋势状态
        chg = h["growth"]
        vol = h["volatility"] or 0
        if chg is not None and chg >= 15 and up >= 2 and last_up: lab = "持续增长"
        elif chg is not None and chg <= -15 and down >= 2 and last_down: lab = "连续下滑"
        elif vol >= 18: lab = "波动较大"
        elif chg is not None and -5 <= chg <= 10 and vol < 12: lab = "高位稳定"
        elif chg is not None and chg < 0 and up >= 1 and last_up: lab = "恢复增长"
        elif chg is not None and chg > 0 and down >= 1 and last_up: lab = "恢复增长"
        elif chg is not None and chg >= 0: lab = "高位稳定"
        else: lab = "波动较大"
        h["trend_label"] = lab
        h["mom7"] = round((tN - t[li-1]) / t[li-1] * 100, 1) if t[li-1] else None  # 最新月环比
        h["trend_series"] = t
    else:
        h["latest_month"] = ""
        h["net"] = h["growth"] = h["mom7"] = h["volatility"] = None
        h["trend_label"] = "数据不足"
        h["trend_series"] = [h["tth"].get(m) for m in MONTHS if h["tth"].get(m) is not None]
    # 活跃率（月表最新月活跃客户 / 标签客户数）
    lm = h.get("latest_month") or "7月"
    hkeys = norm_hosp(h["name"])
    activeX = set()
    for r in y_rows:
        if r["m"] == lm and r["h"] in hkeys:
            activeX.add(r["h"] + "|" + r["c"])
    h["active7"] = len(activeX)
    h["active_rate"] = round(len(activeX) / h["tag_count"] * 100, 1) if h["tag_count"] else None
    # 客户名单匹配数
    h["cust_n"] = len(cust_by_h.get(h["name"], []))
    # 大客户（观念≥治痛管理：客户名单 c_now 主词在治痛管理及以上）
    big = [c for c in cust_by_h.get(h["name"], []) if c["c_now"] in ("治痛管理","资深治痛管理","治痛管理大师")]
    h["big_n"] = len(big)
    h["big_tth7"] = round(sum(c["tth"].get(lm) or 0 for c in big), 1)
    t7v = h["tth"].get(lm)
    if t7v is None and h.get("trend_series"): t7v = h["trend_series"][-1]
    h["big_share"] = round(h["big_tth7"] / t7v * 100, 1) if t7v and h["big_tth7"] else None

# 汇总大区KPI
h7 = sum((h["tth"].get("7月") or 0) for h in hospitals)
h3 = sum((h["tth"].get("3月") or 0) for h in hospitals)
region["kpi"] = {
    "hosp30_t7": round(h7,1), "hosp30_t3": round(h3,1),
    "hosp30_net": round(h7-h3,1),
    "hosp30_growth": round((h7-h3)/h3*100,1) if h3 else None,
    "n_hosp": len(hospitals), "n_top": sum(1 for h in hospitals if h["level"]=="TOP"),
    "n_core": sum(1 for h in hospitals if h["level"]=="CORE"),
    "n_active7_region": region["active_month"].get("7月",0),
    "n_tag": sum((h["tag_count"] or 0) for h in hospitals),
    "n_a7": sum((h["abcd"].get("A") or 0) for h in hospitals),
    "t7_region": region["month_total"].get("7月",0),
    "t6_region": region["month_total"].get("6月",0),
    "mom_region": round((region["month_total"].get("7月",0)-region["month_total"].get("6月",0))/region["month_total"].get("6月",1)*100,1),
}

out = {"meta": {"generated": "2026-08-26", "months": MONTHS},
       "region": region, "hospitals": hospitals,
       "customers": customers, "key_customers": key_customers}
with open(_p("dashboard_data.json"),"w",encoding="utf-8") as f:
    json.dump(out, f, ensure_ascii=False)
print("OK 输出:", len(json.dumps(out,ensure_ascii=False)), "bytes")
print("医院:", len(hospitals), "客户明细:", len(customers), "重点客户:", len(key_customers))
print("大区7月TTH:", region["month_total"].get("7月"), " 30家7月TTH:", h7)
print("观念分布:", {c: region["concept_dist"][c] for c in CONCEPTS})
